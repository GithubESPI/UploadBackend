import base64
import logging
from typing import List
import fitz
from fastapi import APIRouter, UploadFile, File, HTTPException
from fastapi.responses import JSONResponse
import os
import openpyxl
import pypandoc
from docx2pdf import convert
import requests
from app.core.config import settings
from app.services.api_service import fetch_api_data
from app.utils.date_utils import sum_durations, format_minutes_to_duration
from app.services.excel_service import process_excel_file
from docx import Document

# Configure the logger
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

router = APIRouter()

def normalize_title(title):
    import re
    if not isinstance(title, str):
        title = str(title)
    return re.sub(r'\W+', '', title).lower()

async def fetch_api_data_for_template(headers):
    # Fetch API data for students
    api_url = f"{settings.YPAERO_BASE_URL}/r/v1/formation-longue/apprenants?codesPeriode=2"
    api_data = await fetch_api_data(api_url, headers)

    # Fetch API data for groups
    groupes_api_url = f"{settings.YPAERO_BASE_URL}/r/v1/formation-longue/groupes"
    groupes_data = await fetch_api_data(groupes_api_url, headers)

    # Fetch API data for absences
    absences_api_url = f"{settings.YPAERO_BASE_URL}/r/v1/absences/01-01-2023/31-12-2024"
    absences_data = await fetch_api_data(absences_api_url, headers)

    return api_data, groupes_data, absences_data

def extract_appreciations_from_word(word_path):
    try:
        doc = Document(word_path)
        appreciations = {}
        current_name = None
        for table in doc.tables:
            for row in table.rows:
                cells = row.cells
                if len(cells) >= 2:
                    name = cells[0].text.strip()
                    appreciation = cells[1].text.strip()
                    if name and appreciation:
                        appreciations[name] = appreciation
        return appreciations
    except Exception as e:
        logger.error("Failed to extract appreciations from Word document", exc_info=True)
        return {}


def extract_code_apprenant(pdf_path: str) -> str:
    """
    Extract the codeApprenant from the given PDF file.
    Assumes the codeApprenant is a numeric value present in the text.
    """
    with fitz.open(pdf_path) as pdf_document:
        for page_num in range(pdf_document.page_count):
            page = pdf_document.load_page(page_num)
            text = page.get_text("text")
            logger.info(f"Extracted text from {pdf_path}: {text}")  # Log the full extracted text for debugging
            # Extract CodeApprenant by looking for the "Identifiant :" followed by a number
            lines = text.split('\n')
            for line in lines:
                if "Identifiant :" in line:
                    logger.info(f"Identifiant line found: {line}")  # Log the line containing the Identifiant
                    # Extract the identifiant from the line
                    parts = line.split("Identifiant :")
                    if len(parts) > 1:
                        code_apprenant = parts[1].strip()
                        logger.info(f"Extracted code_apprenant before conversion: {code_apprenant}")
                        if code_apprenant.replace('.', '', 1).isdigit():
                            return str(int(float(code_apprenant)))  # Convert to integer
    return None

async def process_file(uploaded_wb, template_path, columns_config):
    template_wb = openpyxl.load_workbook(template_path, data_only=True)
    uploaded_ws = uploaded_wb.active
    template_ws = template_wb.active

    header_row_uploaded = 4
    header_row_template = 1

    uploaded_titles = {normalize_title(uploaded_ws.cell(row=header_row_uploaded, column=col).value): col 
    for col in range(1, uploaded_ws.max_column + 1) 
    if uploaded_ws.cell(row=header_row_uploaded, column=col).value is not None}

    template_titles = {normalize_title(template_ws.cell(row=header_row_template, column=col).value): col 
    for col in range(1, template_ws.max_column + 1) 
    if template_ws.cell(row=header_row_template, column=col).value is not None}

    matching_columns = {uploaded_title: (uploaded_titles[uploaded_title], template_titles[template_title]) 
                        for uploaded_title in uploaded_titles 
                        for template_title in template_titles 
                        if uploaded_title == template_title}

    if not matching_columns:
        return JSONResponse(content={"message": "No matching columns found, leaving new table empty."})

    # Set "Nom" in cell B2 of the template (corrected cell location)
    template_ws.cell(row=header_row_template + 1, column=columns_config['name_column_index_template']).value = "Nom"

    headers = {
        'X-Auth-Token': settings.YPAERO_API_TOKEN,
        'Content-Type': 'application/json'
    }

    api_data, groupes_data, absences_data = await fetch_api_data_for_template(headers)

    # Ensure api_data, groupes_data, and absences_data are dictionaries
    if not isinstance(api_data, dict) or not isinstance(groupes_data, dict) or not isinstance(absences_data, dict):
        raise HTTPException(status_code=500, detail="Unexpected API response format")

    # Debug logging
    logger.debug(f"API Data: {api_data}")
    logger.debug(f"Groupes Data: {groupes_data}")
    logger.debug(f"Absences Data: {absences_data}")

    # Create dictionaries from API data for quick lookup
    api_dict = {normalize_title(apprenant['nomApprenant'] + apprenant['prenomApprenant']): apprenant for key, apprenant in api_data.items()}
    groupes_dict = {groupe['codeGroupe']: groupe for groupe in groupes_data.values()}
    absences_summary = {}
    for absence in absences_data.values():
        apprenant_id = absence.get('codeApprenant')
        duration = int(absence.get('duree', 0))

        if apprenant_id not in absences_summary:
            absences_summary[apprenant_id] = {'justified': [], 'unjustified': [], 'delays': []}

        if absence.get('isJustifie'):
            absences_summary[apprenant_id]['justified'].append(duration)
        elif absence.get('isRetard'):
            absences_summary[apprenant_id]['delays'].append(duration)
        else:
            absences_summary[apprenant_id]['unjustified'].append(duration)

    logger.debug(f"API Dictionary: {api_dict}")
    logger.debug(f"Groupes Dictionary: {groupes_dict}")
    logger.debug(f"Absences Summary: {absences_summary}")

    exclude_phrase = 'moyennedugroupe'  # Normalized exclude phrase
    for row in range(header_row_uploaded + 1, uploaded_ws.max_row + 1):
        if any(exclude_phrase in normalize_title(uploaded_ws.cell(row=row, column=col).value or '') for col in range(1, uploaded_ws.max_column + 1)):
            continue  # Skip rows that contain 'Moyenne du groupe'

        # Copy names from uploaded file to template
        uploaded_name = uploaded_ws.cell(row=row, column=columns_config['name_column_index_uploaded']).value
        template_row = row - header_row_uploaded + header_row_template + 1
        template_ws.cell(row=template_row, column=columns_config['name_column_index_template']).value = uploaded_name

        # Normalize the name for comparison
        normalized_name = normalize_title(uploaded_name)

        # Debug: Check if normalized name is found in the API data
        if (apprenant_info := api_dict.get(normalized_name)):
            logger.debug(f"Found API data for: {uploaded_name}, {apprenant_info}")

            # Add API data to the template if columns exist
            template_ws.cell(row=template_row, column=columns_config['code_apprenant_column_index_template']).value = apprenant_info.get('codeApprenant', 'N/A')
            template_ws.cell(row=template_row, column=columns_config['date_naissance_column_index_template']).value = apprenant_info.get('dateNaissance', 'N/A')
            if 'inscriptions' in apprenant_info and apprenant_info['inscriptions']:
                template_ws.cell(row=template_row, column=columns_config['nom_site_column_index_template']).value = apprenant_info['inscriptions'][0]['site'].get('nomSite', 'N/A')
            
            code_groupe = apprenant_info.get('informationsCourantes', {}).get('codeGroupe', None)
            if code_groupe and code_groupe in groupes_dict:
                groupe_info = groupes_dict[code_groupe]
                template_ws.cell(row=template_row, column=columns_config['code_groupe_column_index_template']).value = groupe_info.get('codeGroupe', 'N/A')
                template_ws.cell(row=template_row, column=columns_config['nom_groupe_column_index_template']).value = groupe_info.get('nomGroupe', 'N/A')
                template_ws.cell(row=template_row, column=columns_config['etendu_groupe_column_index_template']).value = groupe_info.get('etenduGroupe', 'N/A')

            # Add absence data to the template if columns exist
            apprenant_id = apprenant_info.get('codeApprenant')
            abs_info = absences_summary.get(apprenant_id, {'justified': [], 'unjustified': [], 'delays': []})
            
            # Debugging the values before writing them to the sheet
            justified_duration = sum_durations(abs_info['justified']) or "00h00"
            unjustified_duration = sum_durations(abs_info['unjustified']) or "00h00"
            delays_duration = sum_durations(abs_info['delays']) or "00h00"
            logger.debug(f"Row {row} - Justified: {justified_duration}, Unjustified: {unjustified_duration}, Delays: {delays_duration}")

            template_ws.cell(row=template_row, column=columns_config['duree_justifie_column_index_template']).value = justified_duration
            template_ws.cell(row=template_row, column=columns_config['duree_non_justifie_column_index_template']).value = unjustified_duration
            template_ws.cell(row=template_row, column=columns_config['duree_retard_column_index_template']).value = delays_duration

        for uploaded_title, (src_col, dest_col) in matching_columns.items():
            src_cell = uploaded_ws.cell(row=row, column=src_col)
            dest_cell = template_ws.cell(row=template_row, column=dest_col)
            dest_cell.value = src_cell.value

    # Remove duplicates
    for col in range(1, template_ws.max_column + 1):
        if template_ws.cell(row=header_row_template + 1, column=col).value == template_ws.cell(row=header_row_template, column=col).value:
            template_ws.cell(row=header_row_template + 1, column=col).value = None

    # Remove duplicate "Note" in row 3
    for col in range(1, template_ws.max_column + 1):
        if template_ws.cell(row=header_row_template + 2, column=col).value == "Note":
            template_ws.cell(row=header_row_template + 2, column=col).value = None

    # After all processing and before saving, remove the target phrase
    target_phrase = "* Attention, le total des absences prend en compte toutes les absences aux séances sur la période concernée. S'il existe des absences sur des matières qui ne figurent pas dans le relevé, elles seront également comptabilisées."
    for row in template_ws.iter_rows():
        for cell in row:
            if cell.value == target_phrase:
                cell.value = None

    return template_wb

def update_excel_with_appreciations(template_wb, appreciations, columns_config):
    template_ws = template_wb.active
    for row in range(2, template_ws.max_row + 1):  # Start from row 2 to match the data rows
        name = template_ws.cell(row=row, column=columns_config['name_column_index_template']).value
        if name and name in appreciations:
            template_ws.cell(row=row, column=columns_config['appreciation_column_index_template']).value = appreciations[name]
    return template_wb

@router.post("/upload-and-integrate-excel-and-word")
async def upload_and_integrate_excel_and_word(excel_file: UploadFile = File(...), word_file: UploadFile = File(...)):
    try:
        # Save the uploaded files temporarily
        temp_excel_path = os.path.join(settings.UPLOAD_DIR, excel_file.filename)
        temp_word_path = os.path.join(settings.UPLOAD_DIR, word_file.filename)
        with open(temp_excel_path, 'wb') as temp_excel_file:
            temp_excel_file.write(await excel_file.read())
        with open(temp_word_path, 'wb') as temp_word_file:
            temp_word_file.write(await word_file.read())

        # Load the uploaded Excel file
        uploaded_wb = openpyxl.load_workbook(temp_excel_path, data_only=True)
        uploaded_ws = uploaded_wb.active

        # Extract specific cells values for template matching
        uploaded_values = []
        for cell in ['C4', 'F4', 'I4', 'L4', 'O4', 'R4', 'U4', 'X4', 'AA4', 'AD4', 'AG4', 'AJ4', 'AM4', 'AP4', 'AS4', 'AV4', 'AY4', 'BB4', 'BE4', 'BH4']:
            cell_value = uploaded_ws[cell].value
            if cell_value is not None:
                uploaded_values.append(cell_value)

        # Log the extracted values for debugging
        logger.debug(f"Extracted values from uploaded file: {uploaded_values}")

        # Define the template paths
        templates = {
            "MAPI": settings.M1_S1_MAPI_TEMPLATE,
            "MAGI": settings.M1_S1_MAGI_TEMPLATE,
            "MEFIM": settings.M1_S1_MEFIM_TEMPLATE,
            "MAPI_S2": settings.M1_S2_MAPI_TEMPLATE,
            "MAGI_S2": settings.M1_S2_MAGI_TEMPLATE,
            "MEFIM_S2": settings.M1_S2_MEFIM_TEMPLATE,
            "MAPI_S3": settings.M2_S3_MAPI_TEMPLATE,
            "MAGI_S3": settings.M2_S3_MAGI_TEMPLATE,
            "MEFIM_S3": settings.M2_S3_MEFIM_TEMPLATE,
            "MAPI_S4": settings.M2_S4_MAPI_TEMPLATE,
            "MAGI_S4": settings.M2_S4_MAGI_TEMPLATE,
            "MEFIM_S4": settings.M2_S4_MEFIM_TEMPLATE,
        }

        # Define the matching values for each template
        matching_values = {
            "MAPI": ['UE 1 – Economie & Gestion', 'Stratégie et Solutions Immobilières', 'Finance Immobilière', 'Economie Immobilière I', 'UE 2 – Droit', 'Droit des Affaires et des Contrats', 'UE 3 – Aménagement & Urbanisme', 'Ville et Développements Urbains', "Politique de l'Habitat", 'UE 4 – Compétences Professionnalisantes', 'Real Estate English', "Rencontres de l'Immobilier", 'ESPI Career Services', 'ESPI Inside', 'Immersion Professionnelle', 'Projet Voltaire', 'UE SPE – MAPI', 'Etude Foncière', "Montage d'une Opération de Promotion Immobilière", 'Acquisition et Dissociation du Foncier'],
            "MAGI": ['UE 1 – Economie & Gestion', 'Stratégie et Solutions Immobilières', 'Finance Immobilière', 'Economie Immobilière I', 'UE 2 – Droit', 'Droit des Affaires et des Contrats', 'UE 3 – Aménagement & Urbanisme', 'Ville et Développements Urbains', "Politique de l'Habitat", 'UE 4 – Compétences Professionnalisantes', 'Real Estate English', "Rencontres de l'Immobilier", 'ESPI Career Services', 'ESPI Inside', 'Immersion Professionnelle', 'Projet Voltaire', 'UE SPE – MAGI', 'Baux Commerciaux et Gestion Locative', 'Actifs Tertiaires en Copropriété', 'Techniques du Bâtiment'],
            "MEFIM": ['UE 1 – Economie & Gestion', 'Stratégie et Solutions Immobilières', 'Finance Immobilière', 'Economie Immobilière I', 'UE 2 – Droit', 'Droit des Affaires et des Contrats', 'UE 3 – Aménagement & Urbanisme', 'Ville et Développements Urbains', "Politique de l'Habitat", 'UE 4 – Compétences Professionnalisantes', 'Real Estate English', "Rencontres de l'Immobilier", 'ESPI Career Services', 'ESPI Inside', 'Immersion Professionnelle', 'Projet Voltaire', 'UE SPE – MEFIM', "Les Fondamentaux de l'Evaluation", 'Analyse et Financement Immobilier', 'Modélisation Financière'],
            "MAPI_S2": ['UE 1 – Economie & Gestion', "Marketing de l'Immobilier", 'Investissement et Financiarisation', 'Fiscalité', 'UE 2 – Droit', "Droit de l'Urbanisme et de la Construction", "Déontologie en France et à l'International", 'UE 4 – Compétences Professionnalisantes', 'Immersion Professionnelle', 'Real Estate English', 'Atelier Méthodologie de la Recherche', 'Techniques de Négociation', "Rencontres de l'Immobilier", 'ESPI Inside', 'Projet Voltaire', 'UE SPE – MAPI', "Droit de la Promotion Immobilière", "Montage d'une Opération de Logement", 'Financement des Opérations de Promotion Immobilière', "Logement Social et Accession Sociale"],
            "MAGI_S2": ['UE 1 – Economie & Gestion', "Marketing de l'Immobilier", 'Investissement et Financiarisation', 'Fiscalité', 'UE 2 – Droit', "Droit de l'Urbanisme et de la Construction", "Déontologie en France et à l'International", 'UE 4 – Compétences Professionnalisantes', 'Immersion Professionnelle', 'Real Estate English', 'Atelier Méthodologie de la Recherche', 'Techniques de Négociation', "Rencontres de l'Immobilier", 'ESPI Inside', 'Projet Voltaire', 'UE SPE – MAGI', "Budget d'Exploitation et de Travaux", 'Développement et Stratégie Commerciale', 'Technique et Conformité des Immeubles', "Gestion de l'Immobilier - Logistique et Data Center"],
            "MEFIM_S2": ['UE 1 – Economie & Gestion', "Marketing de l'Immobilier", 'Investissement et Financiarisation', 'Fiscalité', 'UE 2 – Droit', "Droit de l'Urbanisme et de la Construction", "Déontologie en France et à l'International", 'UE 4 – Compétences Professionnalisantes', 'Immersion Professionnelle', 'Real Estate English', 'Atelier Méthodologie de la Recherche', 'Techniques de Négociation', "Rencontres de l'Immobilier", 'ESPI Inside', 'Projet Voltaire', 'UE SPE – MEFIM', "Marché d'Actifs Immobiliers", "Baux Commerciaux", 'Evaluation des Actifs Résidentiels', "Audit et Gestion des Immeubles"],
            "MAPI_S3": ['UE 1 – Economie & Gestion', "PropTech et Innovation", 'Economie Immobilière II', 'UE 3 – Aménagement & Urbanisme', "Stratégies et Aménagement des Territoires I", "UE 4 – Compétences Professionnalisantes", 'Communication Digitale, Ecrite et Orale', 'Immersion Professionnelle', 'Real Estate English', 'Méthodologie de la Recherche', "Rencontres de l'Immobilier", 'ESPI Inside', 'UE SPE – MAPI', "Acquisition et Dissociation du Foncier", "Montage des Opérations Tertiaires", "Aménagement et Commande Publique", "Techniques du Bâtiment", "Réhabilitation et Pathologies du Bâtiment"],
            "MAGI_S3": ['UE 1 – Economie & Gestion', "PropTech et Innovation", 'Economie Immobilière II', 'UE 3 – Aménagement & Urbanisme', "Stratégies et Aménagement des Territoires I", "UE 4 – Compétences Professionnalisantes", 'Communication Digitale, Ecrite et Orale', 'Immersion Professionnelle', 'Real Estate English', 'Méthodologie de la Recherche', "Rencontres de l'Immobilier", 'ESPI Inside', 'UE SPE – MAGI', "Rénovation Energétique des Actifs Tertiaires", "Arbitrage, Optimisation et Valorisation des Actifs Tertiaires", 'Maintenance et Facility Management', "Réhabilitation et Pathologies du Bâtiment"],
            "MEFIM_S3": ['UE 1 – Economie & Gestion', "PropTech et Innovation", 'Economie Immobilière II', 'UE 3 – Aménagement & Urbanisme', "Stratégies et Aménagement des Territoires I", "UE 4 – Compétences Professionnalisantes", 'Communication Digitale, Ecrite et Orale', 'Immersion Professionnelle', 'Real Estate English', 'Méthodologie de la Recherche', "Rencontres de l'Immobilier", 'ESPI Inside', 'UE SPE – MEFIM', "Droit des Suretés et de la Transmission", 'Due Diligence', "Evaluation d'Actifs Tertiaires et Industriels", "Gestion de Patrimoine"],
            "MAPI_S4": ['UE 1 – Economie & Gestion', "Economie de l'Environnement", 'UE 3 – Aménagement & Urbanisme', "Normalisation, Labellisation", "Stratégies et Aménagement des Territoires II", 'UE 4 – Compétences Professionnalisantes', 'Real Estate English', 'Mémoire de Recherche', "Rencontres de l'Immobilier", 'ESPI Career Services', 'Immersion Professionnelle', 'UE SPE – MAPI', "Business Game Aménagement et Promotion Immobilière", "Fiscalité et Promotion Immobilière", "Contentieux de l'Urbanisme"],
            "MAGI_S4": ['UE 1 – Economie & Gestion', "Economie de l'Environnement", 'UE 3 – Aménagement & Urbanisme', "Normalisation, Labellisation", "Stratégies et Aménagement des Territoires II", 'UE 4 – Compétences Professionnalisantes', 'Real Estate English', 'Mémoire de Recherche', "Rencontres de l'Immobilier", 'ESPI Career Services', 'Immersion Professionnelle', 'UE SPE – MAGI', "Business Game Property Management", "Gestion des Centres Commerciaux", "Gestion de Contentieux et Recouvrement"],
            "MEFIM_S4": ['UE 1 – Economie & Gestion', "Economie de l'Environnement", 'UE 3 – Aménagement & Urbanisme', "Normalisation, Labellisation", "Stratégies et Aménagement des Territoires II", 'UE 4 – Compétences Professionnalisantes', 'Real Estate English', 'Mémoire de Recherche', "Rencontres de l'Immobilier", 'ESPI Career Services', 'Immersion Professionnelle', 'UE SPE – MEFIM', "Business Game Arbitrage et Stratégies d'Investissement", "Fiscalité du Patrimoine", "Fintech et Blockchain"],
        }

        # Define the column configurations for each template
        column_configs = {
            "MAPI": {
                'name_column_index_uploaded': 2,
                'name_column_index_template': 2,
                'code_apprenant_column_index_template': 1,
                'date_naissance_column_index_template': 23,
                'nom_site_column_index_template': 24,
                'code_groupe_column_index_template': 25,
                'nom_groupe_column_index_template': 26,
                'etendu_groupe_column_index_template': 27,
                'duree_justifie_column_index_template': 28,
                'duree_non_justifie_column_index_template': 29,
                'duree_retard_column_index_template': 30,
                'appreciation_column_index_template': 31
            },
            "MAGI": {
                'name_column_index_uploaded': 2,
                'name_column_index_template': 2,
                'code_apprenant_column_index_template': 1,
                'date_naissance_column_index_template': 23,
                'nom_site_column_index_template': 24,
                'code_groupe_column_index_template': 25,
                'nom_groupe_column_index_template': 26,
                'etendu_groupe_column_index_template': 27,
                'duree_justifie_column_index_template': 28,
                'duree_non_justifie_column_index_template': 29,
                'duree_retard_column_index_template': 30,
                'appreciation_column_index_template': 31
            },
            "MEFIM": {
                'name_column_index_uploaded': 2,
                'name_column_index_template': 2,
                'code_apprenant_column_index_template': 1,
                'date_naissance_column_index_template': 23,
                'nom_site_column_index_template': 24,
                'code_groupe_column_index_template': 25,
                'nom_groupe_column_index_template': 26,
                'etendu_groupe_column_index_template': 27,
                'duree_justifie_column_index_template': 28,
                'duree_non_justifie_column_index_template': 29,
                'duree_retard_column_index_template': 30,
                'appreciation_column_index_template': 31
            },
            "MAPI_S2": {
                'name_column_index_uploaded': 2,
                'name_column_index_template': 2,
                'code_apprenant_column_index_template': 1,
                'date_naissance_column_index_template': 23,
                'nom_site_column_index_template': 24,
                'code_groupe_column_index_template': 25,
                'nom_groupe_column_index_template': 26,
                'etendu_groupe_column_index_template': 27,
                'duree_justifie_column_index_template': 28,
                'duree_non_justifie_column_index_template': 29,
                'duree_retard_column_index_template': 30,
                'appreciation_column_index_template': 31
            },
            "MAGI_S2": {
                'name_column_index_uploaded': 2,
                'name_column_index_template': 2,
                'code_apprenant_column_index_template': 1,
                'date_naissance_column_index_template': 23,
                'nom_site_column_index_template': 24,
                'code_groupe_column_index_template': 25,
                'nom_groupe_column_index_template': 26,
                'etendu_groupe_column_index_template': 27,
                'duree_justifie_column_index_template': 28,
                'duree_non_justifie_column_index_template': 29,
                'duree_retard_column_index_template': 30,
                'appreciation_column_index_template': 31
            },
            "MEFIM_S2": {
                'name_column_index_uploaded': 2,
                'name_column_index_template': 2,
                'code_apprenant_column_index_template': 1,
                'date_naissance_column_index_template': 23,
                'nom_site_column_index_template': 24,
                'code_groupe_column_index_template': 25,
                'nom_groupe_column_index_template': 26,
                'etendu_groupe_column_index_template': 27,
                'duree_justifie_column_index_template': 28,
                'duree_non_justifie_column_index_template': 29,
                'duree_retard_column_index_template': 30,
                'appreciation_column_index_template': 31
            },
            "MAPI_S3": {
                'name_column_index_uploaded': 2,
                'name_column_index_template': 2,
                'code_apprenant_column_index_template': 1,
                'date_naissance_column_index_template': 21,
                'nom_site_column_index_template': 22,
                'code_groupe_column_index_template': 23,
                'nom_groupe_column_index_template': 24,
                'etendu_groupe_column_index_template': 25,
                'duree_justifie_column_index_template': 26,
                'duree_non_justifie_column_index_template': 27,
                'duree_retard_column_index_template': 28,
                'appreciation_column_index_template': 29
            },
            "MAGI_S3": {
                'name_column_index_uploaded': 2,
                'name_column_index_template': 2,
                'code_apprenant_column_index_template': 1,
                'date_naissance_column_index_template': 20,
                'nom_site_column_index_template': 21,
                'code_groupe_column_index_template': 22,
                'nom_groupe_column_index_template': 23,
                'etendu_groupe_column_index_template': 24,
                'duree_justifie_column_index_template': 25,
                'duree_non_justifie_column_index_template': 26,
                'duree_retard_column_index_template': 27,
                'appreciation_column_index_template': 28
            },
            "MEFIM_S3": {
                'name_column_index_uploaded': 2,
                'name_column_index_template': 2,
                'code_apprenant_column_index_template': 1,
                'date_naissance_column_index_template': 20,
                'nom_site_column_index_template': 21,
                'code_groupe_column_index_template': 22,
                'nom_groupe_column_index_template': 23,
                'etendu_groupe_column_index_template': 24,
                'duree_justifie_column_index_template': 25,
                'duree_non_justifie_column_index_template': 26,
                'duree_retard_column_index_template': 27,
                'appreciation_column_index_template': 28
            },
            "MAPI_S4": {
                'name_column_index_uploaded': 2,
                'name_column_index_template': 2,
                'code_apprenant_column_index_template': 1,
                'date_naissance_column_index_template': 18,
                'nom_site_column_index_template': 19,
                'code_groupe_column_index_template': 20,
                'nom_groupe_column_index_template': 21,
                'etendu_groupe_column_index_template': 22,
                'duree_justifie_column_index_template': 23,
                'duree_non_justifie_column_index_template': 24,
                'duree_retard_column_index_template': 25,
                'appreciation_column_index_template': 26
            },
            "MAGI_S4": {
                'name_column_index_uploaded': 2,
                'name_column_index_template': 2,
                'code_apprenant_column_index_template': 1,
                'date_naissance_column_index_template': 18,
                'nom_site_column_index_template': 19,
                'code_groupe_column_index_template': 20,
                'nom_groupe_column_index_template': 21,
                'etendu_groupe_column_index_template': 22,
                'duree_justifie_column_index_template': 23,
                'duree_non_justifie_column_index_template': 24,
                'duree_retard_column_index_template': 25,
                'appreciation_column_index_template': 26
            },
            "MEFIM_S4": {
                'name_column_index_uploaded': 2,
                'name_column_index_template': 2,
                'code_apprenant_column_index_template': 1,
                'date_naissance_column_index_template': 18,
                'nom_site_column_index_template': 19,
                'code_groupe_column_index_template': 20,
                'nom_groupe_column_index_template': 21,
                'etendu_groupe_column_index_template': 22,
                'duree_justifie_column_index_template': 23,
                'duree_non_justifie_column_index_template': 24,
                'duree_retard_column_index_template': 25,
                'appreciation_column_index_template': 26
            },
        }

        template_to_use = None
        columns_config = None
        for template, values in matching_values.items():
            if uploaded_values[:len(values)] == values:
                template_to_use = templates[template]
                columns_config = column_configs[template]
                break

        if not template_to_use or not columns_config:
            raise HTTPException(status_code=400, detail="No matching template found")

        template_wb = await process_file(uploaded_wb, template_to_use, columns_config)

        # Extract appreciations from the uploaded Word file
        appreciations = extract_appreciations_from_word(temp_word_path)
        logger.debug(f"Extracted appreciations: {appreciations}")

        # Update the template with the extracted appreciations if needed
        template_wb = update_excel_with_appreciations(template_wb, appreciations, columns_config)

        # Use the template name for the output file
        template_name = os.path.basename(template_to_use).replace('.xlsx', '')
        output_path = os.path.join(settings.OUTPUT_DIR, f'{template_name}.xlsx')
        template_wb.save(output_path)

    except Exception as e:
        logger.error("Failed to process the file", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    
@router.post("/generate-bulletins")
async def generate_bulletins(file: UploadFile = File(...)):
    try:
        temp_file_path = os.path.join(settings.UPLOAD_DIR, file.filename)
        with open(temp_file_path, 'wb') as temp_file:
            temp_file.write(await file.read())

        output_dir = os.path.join(settings.OUTPUT_DIR, 'bulletins')
        os.makedirs(output_dir, exist_ok=True)

        # Process the Excel file to generate Word bulletins (intermediate step)
        bulletin_paths = process_excel_file(temp_file_path, output_dir)

        # Convert all Word documents in the output_dir to PDF
        convert(output_dir)

        pdf_bulletin_paths = [
            os.path.join(output_dir, filename.replace('.docx', '.pdf'))
            for filename in os.listdir(output_dir)
            if filename.endswith('.pdf')
        ]

        # Optionally, remove the Word documents after conversion
        for bulletin_path in bulletin_paths:
            os.remove(bulletin_path)

        return JSONResponse(content={"message": "Bulletins PDF générés avec succès", "bulletins": pdf_bulletin_paths})

    except Exception as e:
        logger.error("Failed to generate bulletins", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/generate-and-import-bulletins")
async def generate_and_import_bulletins(file: UploadFile = File(...)):
    try:
        # Step 1: Save the uploaded file
        temp_file_path = os.path.join(settings.UPLOAD_DIR, file.filename)
        with open(temp_file_path, 'wb') as temp_file:
            temp_file.write(await file.read())

        # Step 2: Generate PDF bulletins from the Excel file
        output_dir = os.path.join(settings.OUTPUT_DIR, 'bulletins')
        os.makedirs(output_dir, exist_ok=True)
        bulletin_paths = process_excel_file(temp_file_path, output_dir)

        # Convert all Word documents to PDF (Assuming process_excel_file generates Word docs)
        convert(output_dir)
        pdf_bulletin_paths = [
            os.path.join(output_dir, filename.replace('.docx', '.pdf'))
            for filename in os.listdir(output_dir)
            if filename.endswith('.pdf')
        ]

        # Optionally, remove the Word documents after conversion
        for bulletin_path in bulletin_paths:
            os.remove(bulletin_path)

        # Step 3: Import each PDF bulletin to Yparéo
        import_errors = []
        for pdf_bulletin_path in pdf_bulletin_paths:
            try:
                # Extract codeApprenant from the PDF
                code_apprenant = extract_code_apprenant(pdf_bulletin_path)
                if not code_apprenant:
                    raise ValueError(f"codeApprenant not found in {pdf_bulletin_path}")
                
                logger.info(f"Extracted codeApprenant: {code_apprenant} from {pdf_bulletin_path}")

                with open(pdf_bulletin_path, 'rb') as pdf_file:
                    file_content = pdf_file.read()
                    encoded_content = base64.b64encode(file_content).decode('utf-8')

                    # Create the JSON payload
                    payload = {
                        "contenu": encoded_content,
                        "nomDocument": os.path.basename(pdf_bulletin_path),
                        "typeMime": "application/pdf",
                        "extension": "pdf",
                    }

                    # Perform the POST request to import the document
                    endpoint = f"/r/v1/document/apprenant/{code_apprenant}/document?codeRepertoire=1000011"
                    url = f"{settings.YPAERO_BASE_URL}{endpoint}"
                    headers = {
                        "X-Auth-Token": settings.YPAERO_API_TOKEN,
                        "Content-Type": "application/json"
                    }

                    response = requests.post(url, headers=headers, json=payload)
                    
                    if response.status_code != 200:
                        import_errors.append({
                            "file": os.path.basename(pdf_bulletin_path),
                            "status_code": response.status_code,
                            "detail": response.text
                        })

            except Exception as import_exc:
                logger.error(f"Failed to import bulletin {pdf_bulletin_path}", exc_info=True)
                import_errors.append({
                    "file": os.path.basename(pdf_bulletin_path),
                    "error": str(import_exc)
                })

        if import_errors:
            return JSONResponse(content={"message": "Bulletins PDF generated, but some failed to import", "errors": import_errors}, status_code=207)
        else:
            return JSONResponse(content={"message": "Bulletins PDF generated and imported successfully", "bulletins": pdf_bulletin_paths})

    except Exception as e:
        logger.error("Failed to generate and import bulletins", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))